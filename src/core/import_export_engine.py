"""
增强的导入导出引擎
提供高质量的数据交换功能，支持多种格式和高级特性
"""

import json
import csv
import logging
import os
import shutil
import zipfile
from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple, Union, Callable
from dataclasses import dataclass, field
from enum import Enum, auto

from PyQt6.QtCore import QObject, pyqtSignal, QThread, QMutex, QMutexLocker

logger = logging.getLogger(__name__)


class ExportFormat(Enum):
    """导出格式"""
    JSON = "json"
    CSV = "csv"
    EXCEL = "xlsx"
    MARKDOWN = "md"
    BACKUP = "backup"


class ImportMode(Enum):
    """导入模式"""
    REPLACE = auto()    # 替换现有数据
    MERGE = auto()      # 合并数据
    APPEND = auto()     # 追加数据
    UPDATE = auto()     # 更新现有数据


class ConflictAction(Enum):
    """冲突处理动作"""
    SKIP = auto()       # 跳过
    REPLACE = auto()    # 替换
    RENAME = auto()     # 重命名
    MERGE = auto()      # 合并
    ASK_USER = auto()   # 询问用户


@dataclass
class ValidationResult:
    """验证结果"""
    is_valid: bool
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    fixed_issues: List[str] = field(default_factory=list)
    
    @property
    def has_errors(self) -> bool:
        return len(self.errors) > 0
    
    @property
    def has_warnings(self) -> bool:
        return len(self.warnings) > 0


@dataclass
class ImportResult:
    """导入结果"""
    success: bool
    imported_count: int = 0
    skipped_count: int = 0
    error_count: int = 0
    conflicts_resolved: int = 0
    validation_result: Optional[ValidationResult] = None
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    
    @property
    def total_processed(self) -> int:
        return self.imported_count + self.skipped_count + self.error_count


@dataclass
class ExportResult:
    """导出结果"""
    success: bool
    exported_count: int = 0
    file_path: str = ""
    file_size: int = 0
    errors: List[str] = field(default_factory=list)
    
    @property
    def file_size_mb(self) -> float:
        return self.file_size / (1024 * 1024)


class ProgressReporter(QObject):
    """进度报告器"""
    
    progress = pyqtSignal(int)  # 进度百分比
    status = pyqtSignal(str)    # 状态信息
    error = pyqtSignal(str)     # 错误信息
    finished = pyqtSignal()     # 完成信号
    
    def __init__(self):
        super().__init__()
        self.is_cancelled = False
        self.current_progress = 0
    
    def report_progress(self, percentage: int, status: str = ""):
        """报告进度"""
        self.current_progress = max(0, min(100, percentage))
        self.progress.emit(self.current_progress)
        if status:
            self.status.emit(status)
    
    def report_status(self, status: str):
        """报告状态"""
        self.status.emit(status)
    
    def report_error(self, error: str):
        """报告错误"""
        self.error.emit(error)
        logger.error(f"Import/Export error: {error}")
    
    def cancel(self):
        """取消操作"""
        self.is_cancelled = True
    
    def finish(self):
        """完成操作"""
        self.finished.emit()


class DataValidator:
    """数据验证器"""
    
    def __init__(self):
        self.required_fields = {
            'codex_entry': ['id', 'title', 'entry_type'],
            'progression': ['title', 'date'],
            'relationship': ['source', 'target', 'type']
        }
        
        self.field_types = {
            'id': str,
            'title': str,
            'entry_type': str,
            'description': str,
            'is_global': bool,
            'track_references': bool,
            'aliases': list,
            'relationships': list,
            'progression': list
        }
    
    def validate_codex_data(self, data: List[Dict]) -> ValidationResult:
        """验证Codex数据"""
        result = ValidationResult(is_valid=True)
        
        if not isinstance(data, list):
            result.is_valid = False
            result.errors.append("数据必须是列表格式")
            return result
        
        seen_ids = set()
        
        for i, entry in enumerate(data):
            entry_errors = self._validate_codex_entry(entry, i, seen_ids)
            result.errors.extend(entry_errors)
        
        result.is_valid = len(result.errors) == 0
        return result
    
    def _validate_codex_entry(self, entry: Dict, index: int, seen_ids: set) -> List[str]:
        """验证单个Codex条目"""
        errors = []
        
        # 检查必需字段
        for field in self.required_fields['codex_entry']:
            if field not in entry:
                errors.append(f"条目 {index}: 缺少必需字段 '{field}'")
            elif not entry[field]:
                errors.append(f"条目 {index}: 字段 '{field}' 不能为空")
        
        # 检查ID唯一性
        entry_id = entry.get('id')
        if entry_id:
            if entry_id in seen_ids:
                errors.append(f"条目 {index}: ID '{entry_id}' 重复")
            else:
                seen_ids.add(entry_id)
        
        # 检查字段类型
        for field, expected_type in self.field_types.items():
            if field in entry and entry[field] is not None:
                if not isinstance(entry[field], expected_type):
                    errors.append(f"条目 {index}: 字段 '{field}' 类型错误，期望 {expected_type.__name__}")
        
        # 检查entry_type有效性
        valid_types = ['CHARACTER', 'LOCATION', 'OBJECT', 'LORE', 'SUBPLOT', 'OTHER']
        if 'entry_type' in entry and entry['entry_type'] not in valid_types:
            errors.append(f"条目 {index}: entry_type '{entry['entry_type']}' 无效")
        
        return errors
    
    def auto_fix_data(self, data: List[Dict]) -> Tuple[List[Dict], List[str]]:
        """自动修复数据"""
        fixed_data = []
        fix_log = []
        
        for i, entry in enumerate(data):
            fixed_entry = entry.copy()
            
            # 生成缺失的ID
            if not fixed_entry.get('id'):
                fixed_entry['id'] = f"auto_generated_{i}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                fix_log.append(f"条目 {i}: 自动生成ID")
            
            # 设置默认值
            if not fixed_entry.get('title'):
                fixed_entry['title'] = f"未命名条目 {i}"
                fix_log.append(f"条目 {i}: 设置默认标题")
            
            if not fixed_entry.get('entry_type'):
                fixed_entry['entry_type'] = 'OTHER'
                fix_log.append(f"条目 {i}: 设置默认类型")
            
            # 确保字段类型正确
            if 'is_global' in fixed_entry and not isinstance(fixed_entry['is_global'], bool):
                fixed_entry['is_global'] = bool(fixed_entry['is_global'])
                fix_log.append(f"条目 {i}: 修正is_global类型")
            
            if 'aliases' in fixed_entry and not isinstance(fixed_entry['aliases'], list):
                if isinstance(fixed_entry['aliases'], str):
                    fixed_entry['aliases'] = [fixed_entry['aliases']]
                else:
                    fixed_entry['aliases'] = []
                fix_log.append(f"条目 {i}: 修正aliases类型")
            
            fixed_data.append(fixed_entry)
        
        return fixed_data, fix_log


class FormatHandler(ABC):
    """格式处理器抽象基类"""
    
    def __init__(self, progress_reporter: Optional[ProgressReporter] = None):
        self.progress_reporter = progress_reporter
        self.validator = DataValidator()
    
    @abstractmethod
    def can_handle(self, file_path: str) -> bool:
        """检查是否可以处理指定文件"""
        pass
    
    @abstractmethod
    def export_data(self, data: List[Dict], file_path: str, **options) -> ExportResult:
        """导出数据"""
        pass
    
    @abstractmethod
    def import_data(self, file_path: str, **options) -> Tuple[List[Dict], ImportResult]:
        """导入数据"""
        pass
    
    def _report_progress(self, percentage: int, status: str = ""):
        """报告进度"""
        if self.progress_reporter:
            self.progress_reporter.report_progress(percentage, status)
    
    def _report_status(self, status: str):
        """报告状态"""
        if self.progress_reporter:
            self.progress_reporter.report_status(status)


class JSONHandler(FormatHandler):
    """JSON格式处理器"""
    
    def can_handle(self, file_path: str) -> bool:
        return file_path.lower().endswith('.json')
    
    def export_data(self, data: List[Dict], file_path: str, **options) -> ExportResult:
        """导出JSON数据"""
        result = ExportResult(success=False)
        
        try:
            self._report_status("准备导出JSON数据...")
            
            # 准备导出数据
            export_data = {
                'version': '1.0',
                'export_time': datetime.now().isoformat(),
                'total_entries': len(data),
                'entries': data,
                'metadata': options.get('metadata', {})
            }
            
            self._report_progress(30, "写入JSON文件...")
            
            # 写入文件
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2, default=str)
            
            self._report_progress(100, "JSON导出完成")
            
            # 获取文件信息
            file_size = os.path.getsize(file_path)
            
            result.success = True
            result.exported_count = len(data)
            result.file_path = file_path
            result.file_size = file_size
            
        except Exception as e:
            error_msg = f"JSON导出失败: {e}"
            result.errors.append(error_msg)
            logger.error(error_msg)
        
        return result
    
    def import_data(self, file_path: str, **options) -> Tuple[List[Dict], ImportResult]:
        """导入JSON数据"""
        result = ImportResult(success=False)
        data = []
        
        try:
            self._report_status("读取JSON文件...")
            
            with open(file_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            self._report_progress(30, "解析JSON数据...")
            
            # 解析数据结构
            if isinstance(json_data, dict) and 'entries' in json_data:
                # 新格式：包含元数据
                data = json_data['entries']
                logger.info(f"JSON version: {json_data.get('version', 'unknown')}")
            elif isinstance(json_data, list):
                # 简单格式：直接是条目列表
                data = json_data
            else:
                raise ValueError("JSON格式不正确")
            
            self._report_progress(60, "验证数据...")
            
            # 验证数据
            validation_result = self.validator.validate_codex_data(data)
            result.validation_result = validation_result
            
            if not validation_result.is_valid and options.get('auto_fix', False):
                self._report_status("自动修复数据...")
                data, fix_log = self.validator.auto_fix_data(data)
                validation_result.fixed_issues.extend(fix_log)
                
                # 重新验证
                validation_result = self.validator.validate_codex_data(data)
                result.validation_result = validation_result
            
            self._report_progress(100, "JSON导入完成")
            
            result.success = validation_result.is_valid
            result.imported_count = len(data) if result.success else 0
            result.error_count = len(validation_result.errors)
            
        except Exception as e:
            error_msg = f"JSON导入失败: {e}"
            result.errors.append(error_msg)
            logger.error(error_msg)
        
        return data, result


class CSVHandler(FormatHandler):
    """CSV格式处理器"""
    
    def can_handle(self, file_path: str) -> bool:
        return file_path.lower().endswith('.csv')
    
    def export_data(self, data: List[Dict], file_path: str, **options) -> ExportResult:
        """导出CSV数据"""
        result = ExportResult(success=False)
        
        try:
            self._report_status("准备导出CSV数据...")
            
            if not data:
                result.errors.append("没有数据可导出")
                return result
            
            # 确定字段
            all_fields = set()
            for entry in data:
                all_fields.update(entry.keys())
            
            # 排序字段，优先显示重要字段
            priority_fields = ['id', 'title', 'entry_type', 'description', 'is_global']
            ordered_fields = []
            
            for field in priority_fields:
                if field in all_fields:
                    ordered_fields.append(field)
                    all_fields.remove(field)
            
            ordered_fields.extend(sorted(all_fields))
            
            self._report_progress(30, "写入CSV文件...")
            
            # 写入CSV
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:  # 使用BOM确保Excel正确显示中文
                writer = csv.DictWriter(f, fieldnames=ordered_fields)
                writer.writeheader()
                
                for i, entry in enumerate(data):
                    # 处理复杂字段
                    csv_entry = {}
                    for field in ordered_fields:
                        value = entry.get(field, '')
                        if isinstance(value, (list, dict)):
                            csv_entry[field] = json.dumps(value, ensure_ascii=False)
                        else:
                            csv_entry[field] = str(value) if value is not None else ''
                    
                    writer.writerow(csv_entry)
                    
                    # 更新进度
                    progress = 30 + (i / len(data)) * 60
                    self._report_progress(int(progress))
            
            self._report_progress(100, "CSV导出完成")
            
            file_size = os.path.getsize(file_path)
            result.success = True
            result.exported_count = len(data)
            result.file_path = file_path
            result.file_size = file_size
            
        except Exception as e:
            error_msg = f"CSV导出失败: {e}"
            result.errors.append(error_msg)
            logger.error(error_msg)
        
        return result
    
    def import_data(self, file_path: str, **options) -> Tuple[List[Dict], ImportResult]:
        """导入CSV数据"""
        result = ImportResult(success=False)
        data = []
        
        try:
            self._report_status("读取CSV文件...")
            
            # 检测编码
            encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']
            csv_data = None
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        reader = csv.DictReader(f)
                        csv_data = list(reader)
                    break
                except UnicodeDecodeError:
                    continue
            
            if csv_data is None:
                raise ValueError("无法识别CSV文件编码")
            
            self._report_progress(30, "解析CSV数据...")
            
            # 转换数据
            for i, row in enumerate(csv_data):
                entry = {}
                for key, value in row.items():
                    if not key:  # 跳过空字段名
                        continue
                    
                    # 尝试解析JSON字段
                    if value.startswith('[') or value.startswith('{'):
                        try:
                            entry[key] = json.loads(value)
                        except:
                            entry[key] = value
                    else:
                        # 类型转换
                        if key == 'is_global' or key == 'track_references':
                            entry[key] = value.lower() in ('true', '1', 'yes', '是')
                        else:
                            entry[key] = value
                
                data.append(entry)
                
                # 更新进度
                progress = 30 + (i / len(csv_data)) * 40
                self._report_progress(int(progress))
            
            self._report_progress(80, "验证数据...")
            
            # 验证数据
            validation_result = self.validator.validate_codex_data(data)
            result.validation_result = validation_result
            
            if not validation_result.is_valid and options.get('auto_fix', False):
                self._report_status("自动修复数据...")
                data, fix_log = self.validator.auto_fix_data(data)
                validation_result.fixed_issues.extend(fix_log)
                validation_result = self.validator.validate_codex_data(data)
                result.validation_result = validation_result
            
            self._report_progress(100, "CSV导入完成")
            
            result.success = validation_result.is_valid
            result.imported_count = len(data) if result.success else 0
            result.error_count = len(validation_result.errors)
            
        except Exception as e:
            error_msg = f"CSV导入失败: {e}"
            result.errors.append(error_msg)
            logger.error(error_msg)
        
        return data, result


class ExcelHandler(FormatHandler):
    """Excel格式处理器"""
    
    def can_handle(self, file_path: str) -> bool:
        return file_path.lower().endswith(('.xlsx', '.xls'))
    
    def export_data(self, data: List[Dict], file_path: str, **options) -> ExportResult:
        """导出Excel数据"""
        result = ExportResult(success=False)
        
        try:
            # 检查是否安装了openpyxl
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                result.errors.append("需要安装openpyxl库：pip install openpyxl")
                return result
            
            self._report_status("准备导出Excel数据...")
            
            if not data:
                result.errors.append("没有数据可导出")
                return result
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "Codex数据"
            
            self._report_progress(20, "写入Excel标题...")
            
            # 确定字段
            all_fields = set()
            for entry in data:
                all_fields.update(entry.keys())
            
            # 排序字段，优先显示重要字段
            priority_fields = ['id', 'title', 'entry_type', 'description', 'is_global', 'track_references']
            ordered_fields = []
            
            for field in priority_fields:
                if field in all_fields:
                    ordered_fields.append(field)
                    all_fields.remove(field)
            
            ordered_fields.extend(sorted(all_fields))
            
            # 设置标题行样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入标题行
            for col_num, field in enumerate(ordered_fields, 1):
                cell = ws.cell(row=1, column=col_num, value=field)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            self._report_progress(40, "写入Excel数据...")
            
            # 写入数据行
            for row_num, entry in enumerate(data, 2):
                for col_num, field in enumerate(ordered_fields, 1):
                    value = entry.get(field, '')
                    
                    # 处理复杂字段
                    if isinstance(value, (list, dict)):
                        if field == 'aliases' and isinstance(value, list):
                            value = ', '.join(value)
                        else:
                            value = str(value)
                    elif isinstance(value, bool):
                        value = "是" if value else "否"
                    else:
                        value = str(value) if value is not None else ''
                    
                    ws.cell(row=row_num, column=col_num, value=value)
                
                # 更新进度
                progress = 40 + (row_num / len(data)) * 50
                self._report_progress(int(progress))
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            self._report_progress(95, "保存Excel文件...")
            
            # 保存文件
            wb.save(file_path)
            
            self._report_progress(100, "Excel导出完成")
            
            file_size = os.path.getsize(file_path)
            result.success = True
            result.exported_count = len(data)
            result.file_path = file_path
            result.file_size = file_size
            
        except Exception as e:
            error_msg = f"Excel导出失败: {e}"
            result.errors.append(error_msg)
            logger.error(error_msg)
        
        return result
    
    def import_data(self, file_path: str, **options) -> Tuple[List[Dict], ImportResult]:
        """导入Excel数据"""
        result = ImportResult(success=False)
        data = []
        
        try:
            # 检查是否安装了openpyxl
            try:
                import openpyxl
            except ImportError:
                result.errors.append("需要安装openpyxl库：pip install openpyxl")
                return data, result
            
            self._report_status("读取Excel文件...")
            
            # 打开工作簿
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            
            self._report_progress(30, "解析Excel数据...")
            
            # 获取标题行
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
                else:
                    break
            
            if not headers:
                raise ValueError("Excel文件中没有找到标题行")
            
            # 读取数据行
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):  # 跳过空行
                    continue
                
                entry = {}
                for col_num, value in enumerate(row):
                    if col_num >= len(headers):
                        break
                    
                    field = headers[col_num]
                    
                    # 处理特殊字段
                    if field in ['is_global', 'track_references'] and value:
                        entry[field] = str(value).lower() in ('true', '1', 'yes', '是', 'true')
                    elif field == 'aliases' and value:
                        # 将逗号分隔的字符串转为列表
                        entry[field] = [alias.strip() for alias in str(value).split(',') if alias.strip()]
                    else:
                        entry[field] = str(value) if value is not None else ''
                
                data.append(entry)
                
                # 更新进度
                if row_num % 10 == 0:
                    progress = 30 + (row_num / ws.max_row) * 40
                    self._report_progress(int(progress))
            
            wb.close()
            
            self._report_progress(80, "验证数据...")
            
            # 验证数据
            validation_result = self.validator.validate_codex_data(data)
            result.validation_result = validation_result
            
            if not validation_result.is_valid and options.get('auto_fix', False):
                self._report_status("自动修复数据...")
                data, fix_log = self.validator.auto_fix_data(data)
                validation_result.fixed_issues.extend(fix_log)
                validation_result = self.validator.validate_codex_data(data)
                result.validation_result = validation_result
            
            self._report_progress(100, "Excel导入完成")
            
            result.success = validation_result.is_valid
            result.imported_count = len(data) if result.success else 0
            result.error_count = len(validation_result.errors)
            
        except Exception as e:
            error_msg = f"Excel导入失败: {e}"
            result.errors.append(error_msg)
            logger.error(error_msg)
        
        return data, result


class MarkdownHandler(FormatHandler):
    """Markdown格式处理器"""
    
    def can_handle(self, file_path: str) -> bool:
        return file_path.lower().endswith(('.md', '.markdown'))
    
    def export_data(self, data: List[Dict], file_path: str, **options) -> ExportResult:
        """导出Markdown数据"""
        result = ExportResult(success=False)
        
        try:
            self._report_status("准备导出Markdown数据...")
            
            if not data:
                result.errors.append("没有数据可导出")
                return result
            
            # 生成Markdown内容
            md_content = []
            
            # 添加标题
            md_content.append("# Codex 数据导出")
            md_content.append("")
            md_content.append(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            md_content.append(f"总条目数: {len(data)}")
            md_content.append("")
            
            self._report_progress(20, "生成Markdown内容...")
            
            # 按类型分组
            type_groups = {}
            for entry in data:
                entry_type = entry.get('entry_type', 'OTHER')
                if entry_type not in type_groups:
                    type_groups[entry_type] = []
                type_groups[entry_type].append(entry)
            
            # 为每个类型生成章节
            for entry_type, entries in type_groups.items():
                md_content.append(f"## {entry_type}")
                md_content.append("")
                
                for entry in entries:
                    # 条目标题
                    title = entry.get('title', '未命名')
                    md_content.append(f"### {title}")
                    md_content.append("")
                    
                    # 基本信息
                    info_lines = []
                    if entry.get('id'):
                        info_lines.append(f"**ID**: {entry['id']}")
                    if entry.get('is_global'):
                        info_lines.append("**类型**: 全局")
                    if entry.get('track_references'):
                        info_lines.append("**追踪引用**: 是")
                    
                    if info_lines:
                        md_content.extend(info_lines)
                        md_content.append("")
                    
                    # 描述
                    description = entry.get('description', '')
                    if description:
                        md_content.append("**描述**:")
                        md_content.append("")
                        md_content.append(description)
                        md_content.append("")
                    
                    # 别名
                    aliases = entry.get('aliases', [])
                    if aliases:
                        md_content.append("**别名**: " + ", ".join(aliases))
                        md_content.append("")
                    
                    # 关系
                    relationships = entry.get('relationships', [])
                    if relationships:
                        md_content.append("**关系**:")
                        md_content.append("")
                        for rel in relationships:
                            if isinstance(rel, dict):
                                rel_type = rel.get('type', '关联')
                                target = rel.get('target_id', '未知')
                                md_content.append(f"- {rel_type}: {target}")
                        md_content.append("")
                    
                    # 进展
                    progression = entry.get('progression', [])
                    if progression:
                        md_content.append("**进展记录**:")
                        md_content.append("")
                        for prog in progression:
                            if isinstance(prog, dict):
                                prog_title = prog.get('title', '进展')
                                prog_date = prog.get('date', '')
                                md_content.append(f"- **{prog_title}** ({prog_date})")
                                if prog.get('description'):
                                    md_content.append(f"  {prog['description']}")
                        md_content.append("")
                    
                    md_content.append("---")
                    md_content.append("")
                
                # 更新进度
                progress = 20 + (len(md_content) / (len(data) * 10)) * 70
                self._report_progress(int(min(progress, 90)))
            
            self._report_progress(95, "写入Markdown文件...")
            
            # 写入文件
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(md_content))
            
            self._report_progress(100, "Markdown导出完成")
            
            file_size = os.path.getsize(file_path)
            result.success = True
            result.exported_count = len(data)
            result.file_path = file_path
            result.file_size = file_size
            
        except Exception as e:
            error_msg = f"Markdown导出失败: {e}"
            result.errors.append(error_msg)
            logger.error(error_msg)
        
        return result
    
    def import_data(self, file_path: str, **options) -> Tuple[List[Dict], ImportResult]:
        """导入Markdown数据（基础解析）"""
        result = ImportResult(success=False)
        result.errors.append("Markdown导入功能暂未实现，请使用JSON或CSV格式")
        return [], result


class BackupHandler(FormatHandler):
    """备份格式处理器"""
    
    def can_handle(self, file_path: str) -> bool:
        return file_path.lower().endswith('.backup') or file_path.lower().endswith('.zip')
    
    def export_data(self, data: List[Dict], file_path: str, **options) -> ExportResult:
        """创建备份"""
        result = ExportResult(success=False)
        
        try:
            self._report_status("创建备份...")
            
            # 准备备份数据
            backup_data = {
                'version': '1.0',
                'backup_time': datetime.now().isoformat(),
                'app_version': options.get('app_version', '1.0'),
                'total_entries': len(data),
                'codex_entries': data,
                'metadata': {
                    'backup_type': 'full',
                    'compression': True,
                    'includes_relationships': True,
                    'includes_progression': True
                }
            }
            
            # 如果包含数据库路径，也备份数据库文件
            db_path = options.get('database_path')
            config_data = options.get('config_data', {})
            
            self._report_progress(30, "打包备份文件...")
            
            # 创建ZIP备份
            with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as backup_zip:
                # 添加主数据文件
                backup_json = json.dumps(backup_data, ensure_ascii=False, indent=2, default=str)
                backup_zip.writestr('codex_data.json', backup_json)
                
                # 添加配置文件
                if config_data:
                    config_json = json.dumps(config_data, ensure_ascii=False, indent=2, default=str)
                    backup_zip.writestr('config.json', config_json)
                
                # 添加数据库文件（如果存在）
                if db_path and os.path.exists(db_path):
                    backup_zip.write(db_path, 'database.db')
                
                # 添加备份信息文件
                info = {
                    'created_at': datetime.now().isoformat(),
                    'total_entries': len(data),
                    'includes_database': bool(db_path and os.path.exists(db_path)),
                    'includes_config': bool(config_data)
                }
                info_json = json.dumps(info, ensure_ascii=False, indent=2)
                backup_zip.writestr('backup_info.json', info_json)
            
            self._report_progress(100, "备份创建完成")
            
            file_size = os.path.getsize(file_path)
            result.success = True
            result.exported_count = len(data)
            result.file_path = file_path
            result.file_size = file_size
            
        except Exception as e:
            error_msg = f"备份创建失败: {e}"
            result.errors.append(error_msg)
            logger.error(error_msg)
        
        return result
    
    def import_data(self, file_path: str, **options) -> Tuple[List[Dict], ImportResult]:
        """恢复备份"""
        result = ImportResult(success=False)
        data = []
        
        try:
            self._report_status("读取备份文件...")
            
            with zipfile.ZipFile(file_path, 'r') as backup_zip:
                # 检查备份信息
                if 'backup_info.json' in backup_zip.namelist():
                    info_data = json.loads(backup_zip.read('backup_info.json').decode('utf-8'))
                    logger.info(f"Backup created at: {info_data.get('created_at')}")
                
                self._report_progress(30, "解压备份数据...")
                
                # 读取主数据
                if 'codex_data.json' in backup_zip.namelist():
                    backup_data = json.loads(backup_zip.read('codex_data.json').decode('utf-8'))
                    data = backup_data.get('codex_entries', [])
                else:
                    raise ValueError("备份文件中缺少主数据")
                
                # 如果需要，恢复数据库文件
                restore_db = options.get('restore_database', False)
                db_target_path = options.get('database_target_path')
                
                if restore_db and 'database.db' in backup_zip.namelist() and db_target_path:
                    self._report_progress(60, "恢复数据库文件...")
                    
                    # 备份现有数据库
                    if os.path.exists(db_target_path):
                        backup_db_path = f"{db_target_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                        shutil.copy2(db_target_path, backup_db_path)
                        logger.info(f"Current database backed up to: {backup_db_path}")
                    
                    # 恢复数据库
                    with open(db_target_path, 'wb') as db_file:
                        db_file.write(backup_zip.read('database.db'))
                
                self._report_progress(80, "验证备份数据...")
            
            # 验证数据
            validation_result = self.validator.validate_codex_data(data)
            result.validation_result = validation_result
            
            self._report_progress(100, "备份恢复完成")
            
            result.success = validation_result.is_valid
            result.imported_count = len(data) if result.success else 0
            result.error_count = len(validation_result.errors)
            
        except Exception as e:
            error_msg = f"备份恢复失败: {e}"
            result.errors.append(error_msg)
            logger.error(error_msg)
        
        return data, result


class ImportExportEngine(QObject):
    """导入导出引擎核心类"""
    
    # 信号定义
    operationStarted = pyqtSignal(str)  # 操作类型
    operationFinished = pyqtSignal(bool, str)  # 成功状态, 结果消息
    
    def __init__(self, codex_manager=None):
        super().__init__()
        
        self.codex_manager = codex_manager
        self.handlers = {}
        self.progress_reporter = ProgressReporter()
        
        # 注册默认处理器
        self._register_handlers()
        
        logger.info("Import/Export engine initialized")
    
    def _register_handlers(self):
        """注册格式处理器"""
        handlers = [
            JSONHandler(self.progress_reporter),
            CSVHandler(self.progress_reporter),
            ExcelHandler(self.progress_reporter),
            MarkdownHandler(self.progress_reporter),
            BackupHandler(self.progress_reporter)
        ]
        
        # 为每个格式注册对应的处理器
        self.handlers[ExportFormat.JSON] = JSONHandler(self.progress_reporter)
        self.handlers[ExportFormat.CSV] = CSVHandler(self.progress_reporter)
        self.handlers[ExportFormat.EXCEL] = ExcelHandler(self.progress_reporter)
        self.handlers[ExportFormat.MARKDOWN] = MarkdownHandler(self.progress_reporter)
        self.handlers[ExportFormat.BACKUP] = BackupHandler(self.progress_reporter)
    
    def get_supported_formats(self) -> List[ExportFormat]:
        """获取支持的格式"""
        return list(self.handlers.keys())
    
    def export_codex_data(self, file_path: str, export_format: ExportFormat, 
                         **options) -> ExportResult:
        """导出Codex数据"""
        try:
            self.operationStarted.emit("export")
            
            if export_format not in self.handlers:
                raise ValueError(f"不支持的导出格式: {export_format}")
            
            # 获取数据
            if self.codex_manager:
                entries = self.codex_manager.get_all_entries()
                data = [self._entry_to_dict(entry) for entry in entries]
            else:
                data = options.get('data', [])
            
            # 添加额外的导出选项
            export_options = {
                'metadata': {
                    'exported_by': 'AI Novel Editor',
                    'export_time': datetime.now().isoformat(),
                    'total_entries': len(data)
                },
                **options
            }
            
            # 执行导出
            handler = self.handlers[export_format]
            result = handler.export_data(data, file_path, **export_options)
            
            # 发送完成信号
            if result.success:
                message = f"成功导出 {result.exported_count} 个条目到 {file_path}"
                self.operationFinished.emit(True, message)
            else:
                message = f"导出失败: {'; '.join(result.errors)}"
                self.operationFinished.emit(False, message)
            
            return result
            
        except Exception as e:
            error_msg = f"导出操作失败: {e}"
            logger.error(error_msg)
            self.operationFinished.emit(False, error_msg)
            return ExportResult(success=False, errors=[error_msg])
    
    def import_codex_data(self, file_path: str, import_mode: ImportMode = ImportMode.MERGE,
                         **options) -> ImportResult:
        """导入Codex数据"""
        try:
            self.operationStarted.emit("import")
            
            # 根据文件扩展名选择处理器
            handler = None
            for fmt, h in self.handlers.items():
                if h.can_handle(file_path):
                    handler = h
                    break
            
            if not handler:
                raise ValueError(f"不支持的文件格式: {file_path}")
            
            # 执行导入
            data, result = handler.import_data(file_path, **options)
            
            # 如果有Codex管理器且数据有效，则更新数据
            if self.codex_manager and result.success and data:
                import_result = self._import_to_codex_manager(data, import_mode)
                result.imported_count = import_result.imported_count
                result.skipped_count = import_result.skipped_count
                result.conflicts_resolved = import_result.conflicts_resolved
            
            # 发送完成信号
            if result.success:
                message = f"成功导入 {result.imported_count} 个条目"
                if result.skipped_count > 0:
                    message += f"，跳过 {result.skipped_count} 个"
                self.operationFinished.emit(True, message)
            else:
                message = f"导入失败: {'; '.join(result.errors)}"
                self.operationFinished.emit(False, message)
            
            return result
            
        except Exception as e:
            error_msg = f"导入操作失败: {e}"
            logger.error(error_msg)
            self.operationFinished.emit(False, error_msg)
            return ImportResult(success=False, errors=[error_msg])
    
    def create_backup(self, backup_path: str, **options) -> ExportResult:
        """创建完整备份"""
        backup_options = {
            'database_path': options.get('database_path'),
            'config_data': options.get('config_data', {}),
            'app_version': options.get('app_version', '1.0')
        }
        
        return self.export_codex_data(backup_path, ExportFormat.BACKUP, **backup_options)
    
    def restore_backup(self, backup_path: str, **options) -> ImportResult:
        """恢复备份"""
        return self.import_codex_data(backup_path, ImportMode.REPLACE, **options)
    
    def _entry_to_dict(self, entry) -> Dict:
        """将Codex条目转换为字典"""
        return {
            'id': entry.id,
            'title': entry.title,
            'entry_type': entry.entry_type.value,
            'description': entry.description,
            'is_global': entry.is_global,
            'track_references': entry.track_references,
            'aliases': entry.aliases or [],
            'relationships': entry.relationships or [],
            'progression': entry.progression or [],
            'created_at': entry.created_at,
            'updated_at': entry.updated_at,
            'metadata': entry.metadata or {}
        }
    
    def _import_to_codex_manager(self, data: List[Dict], import_mode: ImportMode) -> ImportResult:
        """将数据导入到Codex管理器"""
        result = ImportResult(success=True)
        
        for entry_data in data:
            try:
                # 检查条目是否已存在
                existing_entry = None
                if 'id' in entry_data:
                    existing_entry = self.codex_manager.get_entry(entry_data['id'])
                
                if existing_entry:
                    if import_mode == ImportMode.SKIP:
                        result.skipped_count += 1
                        continue
                    elif import_mode == ImportMode.REPLACE:
                        # 更新现有条目
                        self._update_existing_entry(existing_entry, entry_data)
                        result.imported_count += 1
                        result.conflicts_resolved += 1
                    elif import_mode == ImportMode.MERGE:
                        # 合并数据
                        self._merge_entry_data(existing_entry, entry_data)
                        result.imported_count += 1
                        result.conflicts_resolved += 1
                    else:  # UPDATE
                        # 只更新非空字段
                        self._update_non_empty_fields(existing_entry, entry_data)
                        result.imported_count += 1
                else:
                    # 创建新条目
                    new_entry = self._create_entry_from_dict(entry_data)
                    self.codex_manager.add_entry(new_entry)
                    result.imported_count += 1
                    
            except Exception as e:
                result.error_count += 1
                result.errors.append(f"导入条目失败: {e}")
                logger.error(f"Failed to import entry: {e}")
        
        result.success = result.error_count == 0
        return result
    
    def _create_entry_from_dict(self, data: Dict):
        """从字典创建Codex条目"""
        from core.codex_manager import CodexEntry, CodexEntryType
        
        # 转换entry_type
        entry_type = CodexEntryType.OTHER
        try:
            entry_type = CodexEntryType(data.get('entry_type', 'OTHER'))
        except ValueError:
            logger.warning(f"Invalid entry_type: {data.get('entry_type')}")
        
        return CodexEntry(
            id=data.get('id', ''),
            title=data.get('title', ''),
            entry_type=entry_type,
            description=data.get('description', ''),
            is_global=data.get('is_global', False),
            track_references=data.get('track_references', True),
            aliases=data.get('aliases', []),
            relationships=data.get('relationships', []),
            progression=data.get('progression', []),
            metadata=data.get('metadata', {})
        )
    
    def _update_existing_entry(self, entry, data: Dict):
        """更新现有条目"""
        # 简化实现：直接调用更新方法
        if self.codex_manager:
            self.codex_manager.update_entry(
                entry.id,
                title=data.get('title'),
                description=data.get('description'),
                aliases=data.get('aliases'),
                relationships=data.get('relationships'),
                progression=data.get('progression')
            )
    
    def _merge_entry_data(self, entry, data: Dict):
        """合并条目数据"""
        # 合并别名
        existing_aliases = set(entry.aliases or [])
        new_aliases = set(data.get('aliases', []))
        merged_aliases = list(existing_aliases | new_aliases)
        
        # 合并关系
        existing_relationships = entry.relationships or []
        new_relationships = data.get('relationships', [])
        merged_relationships = existing_relationships + new_relationships
        
        # 合并进展
        existing_progression = entry.progression or []
        new_progression = data.get('progression', [])
        merged_progression = existing_progression + new_progression
        
        # 更新条目
        if self.codex_manager:
            self.codex_manager.update_entry(
                entry.id,
                aliases=merged_aliases,
                relationships=merged_relationships,
                progression=merged_progression
            )
    
    def _update_non_empty_fields(self, entry, data: Dict):
        """只更新非空字段"""
        update_data = {}
        
        for field in ['title', 'description', 'aliases', 'relationships', 'progression']:
            if field in data and data[field]:
                update_data[field] = data[field]
        
        if update_data and self.codex_manager:
            self.codex_manager.update_entry(entry.id, **update_data)